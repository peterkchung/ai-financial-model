# About: Builds a BLANK Excel valuation template (two-stage FCFF DCF). Cells
# the pipeline will populate are left empty with green fill + comments tagging
# the schema field they map to. Tunable analyst assumptions ship with
# reasonable defaults.
#
# Run: uv run --with openpyxl python3 scripts/build_template.py

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Border, Side, Alignment
from openpyxl.comments import Comment
from openpyxl.utils import get_column_letter
from pathlib import Path

OUT = Path(__file__).resolve().parents[1] / "templates" / "valuation_template.xlsx"

# ---------- Styling ----------
HEADER_FONT = Font(bold=True, color="FFFFFF", size=11)
HEADER_FILL = PatternFill("solid", fgColor="1F4E78")
SUBHEAD_FONT = Font(bold=True, size=10)
SUBHEAD_FILL = PatternFill("solid", fgColor="D9E1F2")
INPUT_FILL = PatternFill("solid", fgColor="FFF2CC")     # yellow = analyst tunable
ACTUAL_FILL = PatternFill("solid", fgColor="E2EFDA")    # green = pipeline-populated
COMPUTED_FILL = PatternFill("solid", fgColor="FFFFFF")  # white = formula
THIN = Side(border_style="thin", color="BFBFBF")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)

USD_M = '_($* #,##0_);_($* (#,##0);_($* "-"??_);_(@_)'
PCT = '0.0%'
MULT = '0.00"x"'

def style_header(cell):
    cell.font = HEADER_FONT
    cell.fill = HEADER_FILL
    cell.alignment = Alignment(horizontal="left", vertical="center")
    cell.border = BORDER

def style_input(cell, fmt=None):
    cell.fill = INPUT_FILL
    cell.border = BORDER
    if fmt:
        cell.number_format = fmt

def style_actual(cell, fmt=None, schema_field=None):
    cell.fill = ACTUAL_FILL
    cell.border = BORDER
    if fmt:
        cell.number_format = fmt
    if schema_field:
        cell.comment = Comment(f"Pipeline field: {schema_field}", "template")

def style_computed(cell, fmt=None):
    cell.fill = COMPUTED_FILL
    cell.border = BORDER
    if fmt:
        cell.number_format = fmt

# ---------- Default analyst assumptions ----------
risk_free_rate = 0.042
equity_risk_premium = 0.0475
beta = 1.10
pretax_cost_of_debt = 0.055
marginal_tax_rate = 0.25
revenue_growth_y1_5 = 0.08
target_op_margin = 0.15
year_of_convergence = 5
sales_to_capital = 1.50
terminal_growth = risk_free_rate
terminal_cost_of_capital = 0.075
terminal_roic = 0.12

# ---------- Build workbook ----------
wb = Workbook()

# ============== Sheet 1: Cover ==============
ws = wb.active
assert ws is not None
ws.title = "Cover"
ws.column_dimensions["A"].width = 32
ws.column_dimensions["B"].width = 80

ws["A1"] = "Valuation Template (Blank)"
ws["A1"].font = Font(bold=True, size=16)
ws.merge_cells("A1:B1")

cover_rows = [
    ("Ticker", "", "meta.ticker"),
    ("Company name", "", "meta.company_name"),
    ("Valuation date", "", "meta.valuation_date"),
    ("Reporting currency", "USD, millions unless noted", None),
    ("Source — financials", "", "meta.source"),
    ("", "", None),
    ("Methodology", "Two-stage FCFF DCF. Explicit 10-year forecast → terminal value via stable-growth Gordon. Operating leases treated as debt (post-ASC 842).", None),
    ("", "", None),
    ("Cell color legend", "", None),
    ("  Yellow", "Analyst input — tunable", None),
    ("  Green", "Pipeline-populated (sourced from filings/data room)", None),
    ("  White", "Computed (formula)", None),
    ("", "", None),
    ("Sheets", "", None),
    ("  Inputs", "Tunable assumptions + balance-sheet bridge items", None),
    ("  Historicals", "3 years of P&L + cash flow + capex", None),
    ("  Segments", "Revenue & operating income by reportable segment", None),
    ("  Forecast", "10-year FCFF projection + terminal year", None),
    ("  WACC", "Cost of capital build", None),
    ("  Valuation", "Sum-of-PV → equity value → value per share", None),
]
for i, (k, v, field) in enumerate(cover_rows, start=3):
    ws.cell(row=i, column=1, value=k).font = Font(bold=bool(k and not k.startswith("  ")))
    c = ws.cell(row=i, column=2, value=v)
    c.alignment = Alignment(wrap_text=True, vertical="top")
    if field:
        style_actual(c, schema_field=field)

# ============== Sheet 2: Inputs ==============
ws = wb.create_sheet("Inputs")
ws.column_dimensions["A"].width = 50
ws.column_dimensions["B"].width = 16
ws.column_dimensions["C"].width = 60

ws["A1"] = "Assumption"; ws["B1"] = "Value"; ws["C1"] = "Notes"
for col in "ABC":
    style_header(ws[f"{col}1"])

inputs_rows = [
    ("— Capital structure & cost of capital —", None, None, None, None),
    ("Risk-free rate (10Y UST)", risk_free_rate, "Sourced from macro feed; override to lock.", PCT, "actual:macro.risk_free_rate"),
    ("Equity risk premium (mature US)", equity_risk_premium, "Sourced from industry feed (implied ERP); override to lock.", PCT, "actual:industry.equity_risk_premium"),
    ("Levered beta", beta, "Sourced from industry feed; bottom-up override is appropriate for multi-business issuers.", '0.00', "actual:industry.levered_beta"),
    ("Pre-tax cost of debt", pretax_cost_of_debt, "Sourced from macro feed (BAA corp yield); refine by credit rating.", PCT, "actual:macro.baa_corporate_yield"),
    ("Marginal tax rate", marginal_tax_rate, "Statutory + state for jurisdiction of incorporation.", PCT, "input"),
    ("Effective tax rate (current)", None, "Provision / pre-tax income — pipeline pulls from filing.", PCT, "actual:tax.effective_rate"),
    ("", None, None, None, None),
    ("— Operating story —", None, None, None, None),
    ("Compounded revenue growth, Y1-Y5", revenue_growth_y1_5, "Glide path to terminal. Analyst thesis input.", PCT, "input"),
    ("Target pre-tax operating margin (Y10)", target_op_margin, "Sourced from industry feed; override for thesis upside/downside.", PCT, "actual:industry.pretax_operating_margin"),
    ("Year margins reach target", year_of_convergence, "Linear convergence over this many years.", '0', "input"),
    ("Sales-to-capital ratio (reinvestment)", sales_to_capital, "Sourced from industry feed when available.", MULT, "actual:industry.sales_to_capital"),
    ("", None, None, None, None),
    ("— Terminal-year assumptions —", None, None, None, None),
    ("Terminal growth rate (perpetuity)", terminal_growth, "Capped at risk-free rate. Analyst input.", PCT, "input"),
    ("Terminal cost of capital", terminal_cost_of_capital, "Sourced from industry feed (industry WACC); converges to mature-firm.", PCT, "actual:industry.cost_of_capital"),
    ("Terminal ROIC", terminal_roic, "Sourced from industry feed; sustainable competitive return.", PCT, "actual:industry.return_on_invested_capital"),
    ("", None, None, None, None),
    ("— Balance sheet add-backs (pipeline-populated) —", None, None, None, None),
    ("Cash & cash equivalents", None, "Most recent balance sheet.", USD_M, "actual:bs.cash"),
    ("Marketable securities", None, "Most recent balance sheet.", USD_M, "actual:bs.marketable_securities"),
    ("Long-term debt", None, "Most recent balance sheet.", USD_M, "actual:bs.long_term_debt"),
    ("Long-term lease liabilities (treated as debt)", None, "Post-ASC 842; on balance sheet.", USD_M, "actual:bs.lease_liabilities"),
    ("Shares outstanding (millions)", None, "Most recent count from cover page.", '#,##0.00', "actual:shares.outstanding_m"),
]

INPUT_ROW = {}
for i, row in enumerate(inputs_rows, start=2):
    label, val, note, fmt, kind = row
    ws.cell(row=i, column=1, value=label)
    if val is None and (kind is None):
        ws.cell(row=i, column=1).font = SUBHEAD_FONT
        ws.cell(row=i, column=1).fill = SUBHEAD_FILL
        continue
    cell = ws.cell(row=i, column=2, value=val)
    if kind == "input":
        style_input(cell, fmt)
    elif kind and kind.startswith("actual:"):
        style_actual(cell, fmt, schema_field=kind.split(":",1)[1])
    if note:
        ws.cell(row=i, column=3, value=note).alignment = Alignment(wrap_text=True, vertical="top")
    INPUT_ROW[label] = i

# ============== Sheet 3: Historicals ==============
ws = wb.create_sheet("Historicals")
ws.column_dimensions["A"].width = 36
for col in range(2, 6):
    ws.column_dimensions[get_column_letter(col)].width = 16

ws["A1"] = "Line item ($M)"
ws["B1"] = "FY-2"; ws["C1"] = "FY-1"; ws["D1"] = "FY (latest)"
for col in "ABCD":
    style_header(ws[f"{col}1"])

hist_rows = [
    ("Net sales",                       "pl.net_sales"),
    ("Cost of sales",                   "pl.cost_of_sales"),
    ("Fulfillment",                     "pl.fulfillment"),
    ("Technology and infrastructure",   "pl.tech_and_infra"),
    ("Sales and marketing",             "pl.sales_and_marketing"),
    ("General and administrative",      "pl.general_and_admin"),
    ("Other operating expense",         "pl.other_operating"),
    ("Total operating expenses",        "pl.total_opex"),
    ("Operating income",                "pl.operating_income"),
    ("Interest income",                 "pl.interest_income"),
    ("Interest expense",                "pl.interest_expense"),
    ("Other income (expense), net",     "pl.other_income"),
    ("Income before tax",               "pl.income_before_tax"),
    ("Provision for tax",               "pl.tax_provision"),
    ("Net income",                      "pl.net_income"),
    ("",                                None),
    ("Operating margin",                None),
    ("Capex (PP&E purchases, net)",     "cf.capex"),
    ("Cash from operations",            "cf.cash_from_operations"),
    ("Free cash flow",                  "cf.free_cash_flow"),
]

for i, (label, field) in enumerate(hist_rows, start=2):
    if not label:
        continue
    is_subhead = label in (
        "Total operating expenses","Operating income","Net income",
        "Free cash flow","Operating margin"
    )
    ws.cell(row=i, column=1, value=label).font = SUBHEAD_FONT if is_subhead else Font()
    if label == "Operating margin":
        for j, col in enumerate("BCD", start=0):
            f = f"=IFERROR({col}10/{col}2,\"\")"
            style_computed(ws.cell(row=i, column=2+j, value=f), PCT)
        continue
    for j, col in enumerate("BCD", start=0):
        c = ws.cell(row=i, column=2+j)
        style_actual(c, USD_M, schema_field=f"{field}.{['fy_minus_2','fy_minus_1','fy_latest'][j]}")

# ============== Sheet 4: Segments ==============
ws = wb.create_sheet("Segments")
ws.column_dimensions["A"].width = 28
for col in range(2, 6):
    ws.column_dimensions[get_column_letter(col)].width = 14

ws["A1"] = "Segment"; ws["B1"] = "FY-2"; ws["C1"] = "FY-1"; ws["D1"] = "FY (latest)"
ws["E1"] = "YoY growth"
for col in "ABCDE":
    style_header(ws[f"{col}1"])

ws["A2"] = "Net sales by segment ($M)"
ws["A2"].font = SUBHEAD_FONT; ws["A2"].fill = SUBHEAD_FILL

SEG_SLOTS = 5
seg_rev_start = 3
for k in range(SEG_SLOTS):
    row_idx = seg_rev_start + k
    style_actual(ws.cell(row=row_idx, column=1), schema_field=f"segments[{k}].name")
    for j, col in enumerate("BCD"):
        c = ws.cell(row=row_idx, column=2+j)
        style_actual(c, USD_M, schema_field=f"segments[{k}].revenue.{['fy_minus_2','fy_minus_1','fy_latest'][j]}")
    g = ws.cell(row=row_idx, column=5, value=f"=IFERROR(D{row_idx}/C{row_idx}-1,\"\")")
    style_computed(g, PCT)

total_row = seg_rev_start + SEG_SLOTS
ws.cell(row=total_row, column=1, value="Total").font = SUBHEAD_FONT
for j, col in enumerate("BCD"):
    f = f"=SUM({col}{seg_rev_start}:{col}{total_row-1})"
    c = ws.cell(row=total_row, column=2+j, value=f); style_computed(c, USD_M)
    c.font = SUBHEAD_FONT
ws.cell(row=total_row, column=5, value=f"=IFERROR(D{total_row}/C{total_row}-1,\"\")").number_format = PCT

opi_label_row = total_row + 2
ws.cell(row=opi_label_row, column=1, value="Operating income by segment ($M)").font = SUBHEAD_FONT
ws.cell(row=opi_label_row, column=1).fill = SUBHEAD_FILL

opi_start = opi_label_row + 1
for k in range(SEG_SLOTS):
    row_idx = opi_start + k
    style_computed(ws.cell(row=row_idx, column=1, value=f"=A{seg_rev_start+k}"))
    for j, col in enumerate("BCD"):
        c = ws.cell(row=row_idx, column=2+j)
        style_actual(c, USD_M, schema_field=f"segments[{k}].operating_income.{['fy_minus_2','fy_minus_1','fy_latest'][j]}")

opi_total_row = opi_start + SEG_SLOTS
ws.cell(row=opi_total_row, column=1, value="Total").font = SUBHEAD_FONT
for j, col in enumerate("BCD"):
    f = f"=SUM({col}{opi_start}:{col}{opi_total_row-1})"
    c = ws.cell(row=opi_total_row, column=2+j, value=f); style_computed(c, USD_M)
    c.font = SUBHEAD_FONT

margin_label_row = opi_total_row + 2
ws.cell(row=margin_label_row, column=1, value="Operating margin by segment").font = SUBHEAD_FONT
ws.cell(row=margin_label_row, column=1).fill = SUBHEAD_FILL
for k in range(SEG_SLOTS):
    row_idx = margin_label_row + 1 + k
    style_computed(ws.cell(row=row_idx, column=1, value=f"=A{seg_rev_start+k}"))
    for j, col in enumerate("BCD"):
        f = f"=IFERROR({col}{opi_start+k}/{col}{seg_rev_start+k},\"\")"
        c = ws.cell(row=row_idx, column=2+j, value=f); style_computed(c, PCT)

# ============== Sheet 5: Forecast ==============
ws = wb.create_sheet("Forecast")
ws.column_dimensions["A"].width = 36
for col in range(2, 16):
    ws.column_dimensions[get_column_letter(col)].width = 13

headers = ["Line item ($M)", "Base (FY latest)"] + [f"Y{i}" for i in range(1, 11)] + ["Terminal"]
for j, h in enumerate(headers):
    style_header(ws.cell(row=1, column=1+j, value=h))

def inp(name): return f"Inputs!B{INPUT_ROW[name]}"

# Row 2: Revenue growth rate
ws.cell(row=2, column=1, value="Revenue growth rate")
for col_idx in range(3, 8):
    ws.cell(row=2, column=col_idx, value=f"={inp('Compounded revenue growth, Y1-Y5')}").number_format = PCT
for k, col_idx in enumerate(range(8, 13), start=1):
    f = f"={inp('Compounded revenue growth, Y1-Y5')}-(({inp('Compounded revenue growth, Y1-Y5')}-{inp('Terminal growth rate (perpetuity)')})/5)*{k}"
    ws.cell(row=2, column=col_idx, value=f).number_format = PCT
ws.cell(row=2, column=13, value=f"={inp('Terminal growth rate (perpetuity)')}").number_format = PCT
for c in range(2, 14):
    ws.cell(row=2, column=c).fill = COMPUTED_FILL

# Row 3: Revenues
ws.cell(row=3, column=1, value="Revenues")
style_actual(ws.cell(row=3, column=2), USD_M, schema_field="pl.net_sales.fy_latest")
for col_idx in range(3, 14):
    prev = get_column_letter(col_idx-1); cur = get_column_letter(col_idx)
    style_computed(ws.cell(row=3, column=col_idx, value=f"=IFERROR({prev}3*(1+{cur}2),\"\")"), USD_M)

# Row 4: Operating margin
ws.cell(row=4, column=1, value="Operating margin")
bm = ws.cell(row=4, column=2)
style_actual(bm, PCT, schema_field="pl.operating_margin_base")
bm.comment = Comment("Pipeline computes from operating income / net sales (latest fiscal year).", "template")
for k, col_idx in enumerate(range(3, 13), start=1):
    f = (f"=IFERROR(IF({k}>{inp('Year margins reach target')},{inp('Target pre-tax operating margin (Y10)')},"
         f"$B$4+(({inp('Target pre-tax operating margin (Y10)')}-$B$4)/{inp('Year margins reach target')})*{k}),\"\")")
    ws.cell(row=4, column=col_idx, value=f).number_format = PCT
ws.cell(row=4, column=13, value=f"={inp('Target pre-tax operating margin (Y10)')}").number_format = PCT
for c in range(3, 14):
    ws.cell(row=4, column=c).fill = COMPUTED_FILL

# Row 5: EBIT
ws.cell(row=5, column=1, value="EBIT")
for col_idx in range(2, 14):
    cur = get_column_letter(col_idx)
    style_computed(ws.cell(row=5, column=col_idx, value=f"=IFERROR({cur}3*{cur}4,\"\")"), USD_M)

# Row 6: Tax rate
ws.cell(row=6, column=1, value="Tax rate")
ws.cell(row=6, column=2, value=f"={inp('Effective tax rate (current)')}").number_format = PCT
for k, col_idx in enumerate(range(3, 13), start=1):
    f = (f"=IFERROR({inp('Effective tax rate (current)')}+("
         f"({inp('Marginal tax rate')}-{inp('Effective tax rate (current)')})/10)*{k},\"\")")
    ws.cell(row=6, column=col_idx, value=f).number_format = PCT
ws.cell(row=6, column=13, value=f"={inp('Marginal tax rate')}").number_format = PCT
for c in range(2, 14):
    ws.cell(row=6, column=c).fill = COMPUTED_FILL

# Row 7: EBIT(1-t)
ws.cell(row=7, column=1, value="EBIT(1-t)")
for col_idx in range(2, 14):
    cur = get_column_letter(col_idx)
    style_computed(ws.cell(row=7, column=col_idx, value=f"=IFERROR({cur}5*(1-{cur}6),\"\")"), USD_M)

# Row 8: Reinvestment
ws.cell(row=8, column=1, value="Reinvestment (ΔRev / S2C)")
for col_idx in range(3, 13):
    prev = get_column_letter(col_idx-1); cur = get_column_letter(col_idx)
    f = f"=IFERROR(({cur}3-{prev}3)/{inp('Sales-to-capital ratio (reinvestment)')},\"\")"
    style_computed(ws.cell(row=8, column=col_idx, value=f), USD_M)
f = f"=IFERROR(M7*({inp('Terminal growth rate (perpetuity)')}/{inp('Terminal ROIC')}),\"\")"
style_computed(ws.cell(row=8, column=13, value=f), USD_M)

# Row 9: FCFF
ws.cell(row=9, column=1, value="FCFF").font = SUBHEAD_FONT
for col_idx in range(3, 14):
    cur = get_column_letter(col_idx)
    c = ws.cell(row=9, column=col_idx, value=f"=IFERROR({cur}7-{cur}8,\"\")")
    style_computed(c, USD_M); c.font = SUBHEAD_FONT

# Row 10: WACC
ws.cell(row=10, column=1, value="Cost of capital (WACC)")
for col_idx in range(3, 8):
    ws.cell(row=10, column=col_idx, value="=WACC!B14").number_format = PCT
for k, col_idx in enumerate(range(8, 13), start=1):
    f = f"=WACC!B14-((WACC!B14-{inp('Terminal cost of capital')})/5)*{k}"
    ws.cell(row=10, column=col_idx, value=f).number_format = PCT
ws.cell(row=10, column=13, value=f"={inp('Terminal cost of capital')}").number_format = PCT
for c in range(3, 14):
    ws.cell(row=10, column=c).fill = COMPUTED_FILL

# Row 11: Cumulative discount
ws.cell(row=11, column=1, value="Cumulative discount factor")
ws.cell(row=11, column=3, value="=1/(1+C10)").number_format = '0.0000'
for col_idx in range(4, 13):
    prev = get_column_letter(col_idx-1); cur = get_column_letter(col_idx)
    style_computed(ws.cell(row=11, column=col_idx, value=f"={prev}11*(1/(1+{cur}10))"), '0.0000')

# Row 12: PV(FCFF)
ws.cell(row=12, column=1, value="PV(FCFF)")
for col_idx in range(3, 13):
    cur = get_column_letter(col_idx)
    style_computed(ws.cell(row=12, column=col_idx, value=f"=IFERROR({cur}9*{cur}11,\"\")"), USD_M)

# ============== Sheet 6: WACC ==============
ws = wb.create_sheet("WACC")
ws.column_dimensions["A"].width = 36
ws.column_dimensions["B"].width = 14
ws.column_dimensions["C"].width = 60

ws["A1"] = "WACC build"; ws["B1"] = "Value"; ws["C1"] = "Notes"
for col in "ABC":
    style_header(ws[f"{col}1"])

ws["A2"] = "Cost of equity"; ws["A2"].font = SUBHEAD_FONT; ws["A2"].fill = SUBHEAD_FILL
ws["A3"] = "Risk-free rate"; ws["B3"] = f"={inp('Risk-free rate (10Y UST)')}"; ws["B3"].number_format = PCT
ws["A4"] = "Beta"; ws["B4"] = f"={inp('Levered beta')}"; ws["B4"].number_format = '0.00'
ws["A5"] = "Equity risk premium"; ws["B5"] = f"={inp('Equity risk premium (mature US)')}"; ws["B5"].number_format = PCT
ws["A6"] = "Cost of equity = rf + β × ERP"
ws["B6"] = "=B3+B4*B5"; ws["B6"].number_format = PCT; ws["B6"].font = SUBHEAD_FONT

ws["A8"] = "Cost of debt"; ws["A8"].font = SUBHEAD_FONT; ws["A8"].fill = SUBHEAD_FILL
ws["A9"] = "Pre-tax cost of debt"; ws["B9"] = f"={inp('Pre-tax cost of debt')}"; ws["B9"].number_format = PCT
ws["A10"] = "Marginal tax rate"; ws["B10"] = f"={inp('Marginal tax rate')}"; ws["B10"].number_format = PCT
ws["A11"] = "After-tax cost of debt"
ws["B11"] = "=B9*(1-B10)"; ws["B11"].number_format = PCT

ws["A13"] = "Capital structure"; ws["A13"].font = SUBHEAD_FONT; ws["A13"].fill = SUBHEAD_FILL
ws["A14"] = "WACC"; ws["A14"].font = SUBHEAD_FONT
ws["A15"] = "Market cap (override) — for market WACC"
style_actual(ws["B15"], USD_M, schema_field="market.market_cap_m")
ws["C15"] = "Pipeline-populated. If blank, falls back to 70/30 default split."

total_debt_formula = f"({inp('Long-term debt')}+{inp('Long-term lease liabilities (treated as debt)')})"
mcap = "B15"
ws["B14"] = (
    f"=IF(ISNUMBER({mcap}),"
    f"({mcap}/({mcap}+{total_debt_formula}))*B6 + ({total_debt_formula}/({mcap}+{total_debt_formula}))*B11,"
    f"0.7*B6 + 0.3*B11)"
)
ws["B14"].number_format = PCT; ws["B14"].font = SUBHEAD_FONT
ws["C14"] = "Falls back to 70/30 if market cap blank — flag during validation."

# ============== Sheet 7: Valuation ==============
ws = wb.create_sheet("Valuation")
ws.column_dimensions["A"].width = 40
ws.column_dimensions["B"].width = 16
ws.column_dimensions["C"].width = 60

ws["A1"] = "Valuation output"; ws["B1"] = "Value ($M)"; ws["C1"] = "Notes"
for col in "ABC":
    style_header(ws[f"{col}1"])

ws["A2"] = "Sum of PV(FCFF), Y1-Y10"; ws["B2"] = "=SUM(Forecast!C12:L12)"
style_computed(ws["B2"], USD_M)
ws["A3"] = "Terminal FCFF (Y11)"; ws["B3"] = "=Forecast!M9"
style_computed(ws["B3"], USD_M)
ws["A4"] = "Terminal value (Gordon)"
ws["B4"] = f"=IFERROR(B3/({inp('Terminal cost of capital')}-{inp('Terminal growth rate (perpetuity)')}),\"\")"
style_computed(ws["B4"], USD_M)
ws["A5"] = "PV(Terminal value)"; ws["B5"] = "=IFERROR(B4*Forecast!L11,\"\")"
style_computed(ws["B5"], USD_M)
ws["A6"] = "Enterprise value (operating)"; ws["B6"] = "=B2+B5"
style_computed(ws["B6"], USD_M); ws["B6"].font = SUBHEAD_FONT

ws["A8"] = "Bridge to equity value"; ws["A8"].font = SUBHEAD_FONT; ws["A8"].fill = SUBHEAD_FILL
ws["A9"] = "(-) Long-term debt"; ws["B9"] = f"=-{inp('Long-term debt')}"
style_computed(ws["B9"], USD_M)
ws["A10"] = "(-) Long-term lease liabilities"; ws["B10"] = f"=-{inp('Long-term lease liabilities (treated as debt)')}"
style_computed(ws["B10"], USD_M)
ws["A11"] = "(+) Cash & equivalents"; ws["B11"] = f"={inp('Cash & cash equivalents')}"
style_computed(ws["B11"], USD_M)
ws["A12"] = "(+) Marketable securities"; ws["B12"] = f"={inp('Marketable securities')}"
style_computed(ws["B12"], USD_M)
ws["A13"] = "Equity value"; ws["B13"] = "=B6+B9+B10+B11+B12"
style_computed(ws["B13"], USD_M); ws["B13"].font = SUBHEAD_FONT

ws["A15"] = "Shares outstanding (M)"; ws["B15"] = f"={inp('Shares outstanding (millions)')}"
style_computed(ws["B15"], '#,##0.00')
ws["A16"] = "Implied value per share ($)"; ws["B16"] = "=IFERROR(B13/B15,\"\")"
style_computed(ws["B16"], '$#,##0.00')
ws["B16"].font = Font(bold=True, size=12)
ws["B16"].fill = PatternFill("solid", fgColor="FFD966")

ws["A18"] = "Sanity check vs market"
ws["A18"].font = SUBHEAD_FONT; ws["A18"].fill = SUBHEAD_FILL
ws["A19"] = "Current market price ($/share)"
style_actual(ws["B19"], '$#,##0.00', schema_field="market.price_per_share")
ws["A20"] = "Implied / Market"; ws["B20"] = "=IFERROR(B16/B19,\"\")"
ws["B20"].number_format = '0.0%'

wb.save(OUT)
print(f"Wrote {OUT}")
print(f"  Size: {OUT.stat().st_size/1024:.1f} KB")
print(f"  Sheets: {wb.sheetnames}")
