# About: Populator — the bridge between ExtractedFinancials and the blank
# Excel template. The template carries the field-to-cell mapping inline as
# cell comments (e.g. "Pipeline field: pl.net_sales.fy_latest"); the populator
# scans for those comments, resolves each path against the data, and writes.
# Cells whose fields are None in the data are left empty with a reason-code
# comment — never filled with zeros.

from __future__ import annotations
from pathlib import Path
import re
import shutil
from typing import Any, Optional

from openpyxl import load_workbook
from openpyxl.comments import Comment

from ai_financial_model.schema import ExtractedFinancials

PIPELINE_TAG = re.compile(r"Pipeline field:\s*([\w\.\[\]]+)")


def _resolve(data: Any, path: str) -> Any:
    """Walk a dotted path against a Pydantic model or dict.

    Supports list indexing via `[i]`, e.g. `segments[0].revenue.fy_latest`.
    Returns None if any segment is missing or out-of-range — never raises.
    """
    parts = re.findall(r"[\w_]+|\[\d+\]", path)
    cur: Any = data
    for p in parts:
        if cur is None:
            return None
        if p.startswith("["):
            idx = int(p[1:-1])
            try:
                cur = cur[idx]
            except (IndexError, TypeError):
                return None
        else:
            cur = getattr(cur, p, None) if not isinstance(cur, dict) else cur.get(p)
    return cur


def populate_template(
    extracted: ExtractedFinancials,
    template_path: Path,
    output_path: Path,
    provenance: Optional[dict[str, str]] = None,
) -> dict[str, Any]:
    """Copy the blank template to `output_path`, then write all resolvable
    fields into their tagged cells.

    Returns a report dict with summary counts and a per-cell trail suitable
    for serialization to audit.json. The trail entries record the schema
    field, the resolved value (or None), the source ingester (looked up from
    `provenance`, when supplied), and the cell coordinate.

    Output preserves the template's formulas, formatting, and color coding.
    """
    output_path.parent.mkdir(parents=True, exist_ok=True)
    shutil.copy(template_path, output_path)

    provenance = provenance or {}
    wb = load_workbook(output_path)
    populated = 0
    skipped_missing = 0
    default_kept = 0
    seen = 0
    cells: list[dict[str, Any]] = []

    for ws in wb.worksheets:
        for row in ws.iter_rows():
            for cell in row:
                if cell.comment is None:
                    continue
                m = PIPELINE_TAG.search(cell.comment.text)
                if not m:
                    continue
                seen += 1
                path = m.group(1)
                value = _resolve(extracted, path)
                source = provenance.get(path)

                if value is None:
                    if cell.value is None:
                        skipped_missing += 1
                        status = "no_value_extracted"
                        cell.comment = Comment(
                            f"Pipeline field: {path}\nReason: NO_VALUE_EXTRACTED",
                            "pipeline",
                        )
                    else:
                        default_kept += 1
                        status = "default_kept"
                        cell.comment = Comment(
                            f"Pipeline field: {path}\nNo extraction; template default retained.",
                            "pipeline",
                        )
                    cells.append({
                        "sheet": ws.title,
                        "cell": cell.coordinate,
                        "schema_field": path,
                        "value": cell.value,
                        "source": source,
                        "status": status,
                    })
                    continue

                cell.value = value
                cell.comment = Comment(
                    f"Pipeline field: {path}\nSource: {source or extracted.meta.source or 'unknown'}",
                    "pipeline",
                )
                populated += 1
                cells.append({
                    "sheet": ws.title,
                    "cell": cell.coordinate,
                    "schema_field": path,
                    "value": value,
                    "source": source or extracted.meta.source,
                    "status": "populated",
                })

    wb.save(output_path)
    return {
        "tagged_cells": seen,
        "populated": populated,
        "default_kept": default_kept,
        "skipped_missing": skipped_missing,
        "cells": cells,
    }
