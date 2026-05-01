# About: Per-run mapping.md generator. Walks the blank template's tagged
# cells, looks up each schema field's source from the orchestrator's
# provenance map, and produces a Markdown blueprint of how data flows from
# configured ingesters → schema → cells. Lives in coverage/<ticker>/outputs/<ts>/
# alongside the populated model.

from __future__ import annotations
import re
from collections import defaultdict
from datetime import datetime, timezone
from pathlib import Path
from typing import Any, Optional

from openpyxl import load_workbook

from ai_financial_model.generation.populator import PIPELINE_TAG
from ai_financial_model.pipeline import _walk_paths
from ai_financial_model.schema import ExtractedFinancials


def enumerate_template_cells(template_path: Path) -> list[tuple[str, str, str]]:
    """Walk the blank template; return [(sheet, cell_coord, schema_field), ...]
    for every cell carrying a `Pipeline field:` comment."""
    wb = load_workbook(template_path)
    out: list[tuple[str, str, str]] = []
    for ws in wb.worksheets:
        for row in ws.iter_rows():
            for cell in row:
                if cell.comment is None:
                    continue
                m = PIPELINE_TAG.search(cell.comment.text)
                if not m:
                    continue
                out.append((ws.title, cell.coordinate, m.group(1)))
    return out


def build_mapping_md(
    *,
    company_config: dict,
    company_config_path: Optional[Path] = None,
    template_path: Path,
    provenance: dict[str, str],
    extracted: ExtractedFinancials,
) -> str:
    """Compose the markdown document. Pure; doesn't write to disk."""
    meta = company_config.get("meta", {}) or {}
    ticker = meta.get("ticker", "?")
    name = meta.get("company_name", "?")
    val_date = meta.get("valuation_date", "?")
    cik = extracted.meta.cik or "?"
    generated_at = datetime.now(timezone.utc).strftime("%Y-%m-%dT%H-%M-%SZ")

    lines: list[str] = []

    # 1. Header
    lines.append(f"# Valuation mapping — {ticker}")
    lines.append("")
    lines.append(f"- **Company:** {name}")
    lines.append(f"- **Ticker:** {ticker}")
    lines.append(f"- **CIK:** {cik}")
    lines.append(f"- **Valuation date:** {val_date}")
    lines.append(f"- **Generated:** {generated_at} (UTC)")
    if company_config_path is not None:
        lines.append(f"- **Config:** `{company_config_path}`")
    lines.append("")
    lines.append(
        "This document describes for this run: which data sources were "
        "configured, which schema fields they populated, and which template "
        "cells received each value. The timestamp directory + the **Configured "
        "data sources** table below is the record of *what files were used at "
        "what time*."
    )
    lines.append("")

    # 2. Configured data sources
    lines.append("## Configured data sources")
    lines.append("")
    ingesters = company_config.get("ingesters", []) or []
    if ingesters:
        lines.append("| Ingester | Args |")
        lines.append("|---|---|")
        for spec in ingesters:
            t = spec.get("type", "?")
            args = spec.get("args", {}) or {}
            args_str = ", ".join(f"`{k}={v}`" for k, v in args.items())
            lines.append(f"| `{t}` | {args_str or '_(none)_'} |")
    else:
        lines.append("_No ingesters configured._")
    lines.append("")
    if "industry" in (company_config or {}):
        lines.append("**Inline `industry:` block** (analyst calibration, applied last):")
        lines.append("")
        lines.append("| Field | Value |")
        lines.append("|---|---|")
        for k, v in (company_config["industry"] or {}).items():
            lines.append(f"| `{k}` | {v} |")
        lines.append("")

    # 3. Field plan — by template cell
    lines.append("## Field plan — by template cell")
    lines.append("")
    lines.append(
        "For every tagged cell in the template, the schema field it expects "
        "and the source string of the ingester that contributed the value. "
        "Cells with no source are flagged `(unfilled)` — analyst follow-up."
    )
    lines.append("")
    cells = enumerate_template_cells(template_path)
    by_sheet: dict[str, list[tuple[str, str, Optional[str]]]] = defaultdict(list)
    unfilled: list[tuple[str, str, str]] = []
    for sheet, coord, path in cells:
        source = provenance.get(path)
        by_sheet[sheet].append((coord, path, source))
        if source is None:
            unfilled.append((sheet, coord, path))

    for sheet in sorted(by_sheet):
        lines.append(f"### `{sheet}` sheet")
        lines.append("")
        lines.append("| Cell | Schema field | Source |")
        lines.append("|---|---|---|")
        rows = sorted(by_sheet[sheet], key=_sheet_sort_key)
        for coord, path, source in rows:
            src = f"`{source}`" if source else "_(unfilled)_"
            lines.append(f"| `{coord}` | `{path}` | {src} |")
        lines.append("")

    # 4. Unfilled cells
    lines.append("## Unfilled cells")
    lines.append("")
    if unfilled:
        lines.append(
            "These template cells expect a value but no configured source "
            "provided one. Either fill them manually post-run, or configure "
            "an additional ingester."
        )
        lines.append("")
        lines.append("| Sheet | Cell | Schema field |")
        lines.append("|---|---|---|")
        for sheet, coord, path in sorted(unfilled):
            lines.append(f"| {sheet} | `{coord}` | `{path}` |")
    else:
        lines.append("_All tagged cells received a source._")
    lines.append("")

    # 5. Schema fields without template cells
    lines.append("## Schema fields without template cells")
    lines.append("")
    lines.append(
        "These fields were extracted by ingesters but have no corresponding "
        "tagged cell in the template — the data is captured but not displayed "
        "in the model. Add a tagged cell via `scripts/build_template.py` if "
        "you want to surface them."
    )
    lines.append("")
    template_paths = {path for _, _, path in cells}
    extracted_paths = {p for p, _ in _walk_paths(extracted)}
    unused = sorted(extracted_paths - template_paths)
    if unused:
        lines.append("| Schema field | Source |")
        lines.append("|---|---|")
        for path in unused:
            src = provenance.get(path) or extracted.meta.source or "?"
            lines.append(f"| `{path}` | `{src}` |")
    else:
        lines.append("_Every extracted field is mapped to a template cell._")
    lines.append("")

    return "\n".join(lines)


def _sheet_sort_key(row: tuple[Any, ...]) -> tuple:
    """Sort cells by (column letters, row number) for readability."""
    coord = row[0]
    m = re.match(r"^([A-Z]+)(\d+)$", str(coord))
    if not m:
        return (coord, 0)
    return (m.group(1), int(m.group(2)))


def write_mapping(out_dir: Path, **kwargs) -> Path:
    """Convenience: write `out_dir / 'mapping.md'`."""
    out_dir.mkdir(parents=True, exist_ok=True)
    path = out_dir / "mapping.md"
    path.write_text(build_mapping_md(**kwargs))
    return path
