# About: Internal-consistency checks for a populated workbook. Pulls historical
# line items off the Historicals sheet and verifies mechanical ties (total opex,
# operating income, income before tax). Cross-source reconciliation is deferred
# until the pipeline ingests more than one source per company.

from __future__ import annotations
import re
from pathlib import Path
from typing import Optional

from openpyxl import load_workbook

from ai_financial_model.validation.report import (
    Finding, Severity, ValidationReport,
)

# Default tolerance for mechanical ties (variance %).
DEFAULT_TOLERANCE_PCT = 0.5


def _f(val) -> Optional[float]:
    if val is None or isinstance(val, str):
        return None
    try:
        return float(val)
    except (TypeError, ValueError):
        return None


def _variance_pct(expected: float, actual: float) -> float:
    if expected == 0:
        return abs(actual) * 100
    return abs(actual - expected) / abs(expected) * 100


def _check_sum(report: ValidationReport, *, name: str, components: list[Optional[float]],
               total: Optional[float], tolerance_pct: float, refs: list[str]) -> None:
    if total is None or any(c is None for c in components):
        report.findings.append(Finding(
            check=name,
            severity=Severity.YELLOW,
            message="Skipped — one or more inputs missing.",
            cell_refs=refs,
        ))
        return
    expected = sum(c for c in components if c is not None)
    var = _variance_pct(expected, total)
    if var > tolerance_pct:
        report.findings.append(Finding(
            check=name,
            severity=Severity.RED,
            message=f"Variance {var:.2f}% exceeds tolerance {tolerance_pct}%.",
            expected=expected, actual=total, variance_pct=var, cell_refs=refs,
        ))
    else:
        report.findings.append(Finding(
            check=name, severity=Severity.GREEN,
            message=f"OK ({var:.2f}% variance).",
            expected=expected, actual=total, variance_pct=var, cell_refs=refs,
        ))


def validate_workbook(path: Path, tolerance_pct: float = DEFAULT_TOLERANCE_PCT) -> ValidationReport:
    """Run mechanical-tie checks on a populated workbook."""
    wb = load_workbook(path, data_only=False)
    report = ValidationReport()

    if "Historicals" not in wb.sheetnames:
        report.findings.append(Finding(
            check="historicals_present", severity=Severity.RED,
            message="Workbook missing Historicals sheet."))
        return report

    h = wb["Historicals"]

    # Build a label -> {fy_minus_2, fy_minus_1, fy_latest} map.
    rows: dict[str, dict[str, Optional[float]]] = {}
    for row in h.iter_rows(min_row=2, values_only=False):
        label = row[0].value
        if not label:
            continue
        rows[str(label).strip()] = {
            "fy_minus_2": _f(row[1].value),
            "fy_minus_1": _f(row[2].value),
            "fy_latest": _f(row[3].value),
        }

    # Check: Total opex = sum of opex line items, for the latest year.
    opex_components = [
        rows.get("Cost of sales", {}).get("fy_latest"),
        rows.get("Fulfillment", {}).get("fy_latest"),
        rows.get("Technology and infrastructure", {}).get("fy_latest"),
        rows.get("Sales and marketing", {}).get("fy_latest"),
        rows.get("General and administrative", {}).get("fy_latest"),
        rows.get("Other operating expense", {}).get("fy_latest"),
    ]
    _check_sum(
        report,
        name="opex_components_sum_to_total",
        components=opex_components,
        total=rows.get("Total operating expenses", {}).get("fy_latest"),
        tolerance_pct=tolerance_pct,
        refs=["Historicals!B2:D9"],
    )

    # Check: Operating income = Net sales - Total opex.
    ns = rows.get("Net sales", {}).get("fy_latest")
    topx = rows.get("Total operating expenses", {}).get("fy_latest")
    oi = rows.get("Operating income", {}).get("fy_latest")
    if ns is not None and topx is not None and oi is not None:
        expected_oi = ns - topx
        var = _variance_pct(expected_oi, oi)
        sev = Severity.GREEN if var <= tolerance_pct else Severity.RED
        report.findings.append(Finding(
            check="operating_income_tie",
            severity=sev,
            message=("OK" if sev == Severity.GREEN else
                     f"Variance {var:.2f}% exceeds tolerance {tolerance_pct}%."),
            expected=expected_oi, actual=oi, variance_pct=var,
            cell_refs=["Historicals!D10"],
        ))

    # Check: Income before tax = OI + interest income - interest expense + other.
    ii = rows.get("Interest income", {}).get("fy_latest")
    ie = rows.get("Interest expense", {}).get("fy_latest")
    other = rows.get("Other income (expense), net", {}).get("fy_latest")
    ibt = rows.get("Income before tax", {}).get("fy_latest")
    if all(v is not None for v in (oi, ii, ie, other, ibt)):
        # Interest expense in the template is stored as negative.
        expected_ibt = oi + ii + ie + other  # type: ignore[operator]
        var = _variance_pct(expected_ibt, ibt)  # type: ignore[arg-type]
        sev = Severity.GREEN if var <= tolerance_pct else Severity.RED
        report.findings.append(Finding(
            check="income_before_tax_tie",
            severity=sev,
            message=("OK" if sev == Severity.GREEN else
                     f"Variance {var:.2f}% exceeds tolerance {tolerance_pct}%."),
            expected=expected_ibt, actual=ibt, variance_pct=var,
            cell_refs=["Historicals!D14"],
        ))

    # Check: Segment revenue sums tie back to total Net sales.
    # We sum the individual segment rows directly (openpyxl can't evaluate
    # the workbook's SUM formula without a saved cached value).
    if "Segments" in wb.sheetnames:
        seg = wb["Segments"]
        seg_total = 0.0
        any_seen = False
        for row in seg.iter_rows(min_row=3, max_row=7, values_only=False):
            v = _f(row[3].value)
            if v is not None:
                seg_total += v
                any_seen = True
        if any_seen and ns is not None:
            var = _variance_pct(ns, seg_total)
            sev = Severity.GREEN if var <= tolerance_pct else (
                Severity.YELLOW if var <= tolerance_pct * 5 else Severity.RED)
            report.findings.append(Finding(
                check="segments_sum_to_net_sales",
                severity=sev,
                message=(f"Segment total ${seg_total:,.0f}M vs Net sales ${ns:,.0f}M "
                         f"(variance {var:.2f}%)."),
                expected=ns, actual=seg_total, variance_pct=var,
                cell_refs=["Segments!D8", "Historicals!D2"],
            ))

    # Check: Y1 forecast revenue (annualized to a quarter) consistent with
    # the company's own forward guidance for the next quarter, when available.
    # Rough check — Q2 alone isn't representative of a full year, so we use a
    # 10% buffer around the guidance midpoint for the YELLOW band.
    _check_forecast_vs_guidance(report, wb, base_revenue=ns)

    # Check: non-recurring items as a proportion of pre-tax income. If the
    # ingester populated any items, sum them and compare. Skipped silently if
    # the non_recurring_items ingester wasn't enabled.
    _check_non_recurring_proportion(report, wb,
                                    pretax=rows.get("Income before tax", {}).get("fy_latest"))

    return report


def _check_forecast_vs_guidance(report: ValidationReport, wb, *,
                                 base_revenue: Optional[float]) -> None:
    """Compare Y1 forecast quarterly average to forward-guidance range.

    Reads guidance from the Cover sheet (forward_guidance.revenue_low/high)
    and computes Y1 revenue manually as base × (1 + Y1-Y5 growth) since the
    Forecast sheet's Y1 cell is a formula and openpyxl can't evaluate it.
    Skips silently if any input is missing.
    """
    if "Cover" not in wb.sheetnames or base_revenue is None:
        return
    cover = wb["Cover"]
    rev_low: Optional[float] = None
    rev_high: Optional[float] = None
    for row in cover.iter_rows(min_row=1, max_row=cover.max_row, values_only=False):
        if not row or not row[0].value:
            continue
        label = str(row[0].value).strip().lower()
        if label == "revenue, low ($m)":
            rev_low = _f(row[1].value)
        elif label == "revenue, high ($m)":
            rev_high = _f(row[1].value)
    if rev_low is None or rev_high is None:
        return

    # Y1 growth rate from Inputs sheet
    if "Inputs" not in wb.sheetnames:
        return
    growth: Optional[float] = None
    for row in wb["Inputs"].iter_rows(min_row=1, values_only=False):
        if not row or not row[0].value:
            continue
        if str(row[0].value).strip().lower().startswith("compounded revenue growth"):
            growth = _f(row[1].value)
            break
    if growth is None:
        return

    y1_revenue = base_revenue * (1 + growth)
    quarterly_avg = y1_revenue / 4
    midpoint = (rev_low + rev_high) / 2

    if rev_low <= quarterly_avg <= rev_high:
        sev = Severity.GREEN
        msg = (f"Y1 implied Q-avg ${quarterly_avg:,.0f}M is within guidance "
               f"range ${rev_low:,.0f}M–${rev_high:,.0f}M.")
    else:
        var = _variance_pct(midpoint, quarterly_avg)
        # 10% seasonality buffer — Q2 isn't representative of full year
        if var <= 10:
            sev = Severity.YELLOW
            msg = (f"Y1 implied Q-avg ${quarterly_avg:,.0f}M is outside guidance "
                   f"range (${rev_low:,.0f}M–${rev_high:,.0f}M) but within 10% of "
                   f"midpoint ({var:.1f}%). Acceptable given Q2-vs-annual seasonality.")
        else:
            sev = Severity.RED
            msg = (f"Y1 implied Q-avg ${quarterly_avg:,.0f}M deviates {var:.1f}% from "
                   f"guidance midpoint ${midpoint:,.0f}M (range ${rev_low:,.0f}M–"
                   f"${rev_high:,.0f}M). Reconcile growth assumption or investigate "
                   f"seasonality.")
    report.findings.append(Finding(
        check="forecast_y1_consistent_with_guidance",
        severity=sev,
        message=msg,
        expected=midpoint,
        actual=quarterly_avg,
        cell_refs=["Cover!B11", "Cover!B12", "Forecast!C3", "Inputs!B11"],
    ))


_NON_RECURRING_AMOUNT_TAG = re.compile(r"non_recurring_items\[\d+\]\.amount")


def _check_non_recurring_proportion(report: ValidationReport, wb, *,
                                     pretax: Optional[float]) -> None:
    """Sum non_recurring_items[*].amount cells from Cover; compare to pre-tax
    income from Historicals. Severity:
        GREEN  if |sum| / |pretax| < 5%   (immaterial)
        YELLOW if 5% - 15%                (normalize before forecasting)
        RED    if > 15%                   (do NOT capitalize aggregate Other
                                           income into terminal value)
    Skipped silently if no items were populated or if pretax is unavailable.
    """
    if "Cover" not in wb.sheetnames or pretax is None or pretax == 0:
        return

    cover = wb["Cover"]
    total = 0.0
    any_seen = False
    for row in cover.iter_rows():
        for cell in row:
            if cell.comment is None:
                continue
            if not _NON_RECURRING_AMOUNT_TAG.search(cell.comment.text):
                continue
            v = _f(cell.value)
            if v is not None:
                total += v
                any_seen = True
    if not any_seen:
        return  # ingester wasn't enabled; nothing to check

    proportion = abs(total) / abs(pretax)
    if proportion < 0.05:
        sev = Severity.GREEN
        msg = (f"Non-recurring items ${total:,.0f}M = {proportion:.1%} of "
               f"pre-tax income — immaterial.")
    elif proportion < 0.15:
        sev = Severity.YELLOW
        msg = (f"Non-recurring items ${total:,.0f}M = {proportion:.1%} of "
               f"pre-tax income. Normalization recommended before forecasting.")
    else:
        sev = Severity.RED
        msg = (f"Non-recurring items ${total:,.0f}M = {proportion:.1%} of "
               f"pre-tax income exceeds 15% threshold. Do not capitalize "
               f"aggregate Other income into terminal value without explicit "
               f"normalization.")
    report.findings.append(Finding(
        check="non_recurring_items_proportion",
        severity=sev,
        message=msg,
        expected=pretax,
        actual=total,
        cell_refs=["Cover (non-recurring items block)", "Historicals!D14"],
    ))
